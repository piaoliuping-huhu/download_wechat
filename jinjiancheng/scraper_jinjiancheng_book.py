"""
金渐成推荐书单爬虫
从 https://www.jinjiancheng.com/booklist 抓取所有推荐书单，
生成格式化的 Excel 文件，包含分类、书名、作者、封面图、推荐理由等信息。
"""
import requests
import re
import os
import html as html_lib
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def fetch_booklist_html():
    """抓取书单页面 HTML"""
    url = "https://www.jinjiancheng.com/booklist"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    resp = requests.get(url, headers=headers, timeout=30)
    resp.raise_for_status()
    return resp.text


def _extract_book_info(part):
    """从分割后的 part 中提取单本书的完整信息（书名、作者、推荐理由、封面图、详情链接）"""
    # 提取书名 (part 以 "书名》\"," 或类似格式开头)
    title_match = re.search(r'^([^\\]+)》', part)
    if title_match:
        title = f"《{title_match.group(1)}》"
    else:
        # h3 格式: 书名可能不在开头
        h3_match = re.search(r'\\"h3\\".*?\\"children\\":\\"(《[^》]+》)\\"', part)
        if h3_match:
            title = h3_match.group(1)
        else:
            return None

    # 提取作者 - 在 mt-0.5 之后找 children\\":\\" 后面的值
    author = ""
    idx_mt05 = part.find('mt-0.5')
    if idx_mt05 >= 0:
        idx_children = part.find('children', idx_mt05)
        if idx_children >= 0:
            colon_quote = part.find('\\":\\"', idx_children)
            if colon_quote >= 0:
                val_start = colon_quote + 5
                val_end = part.find('\\"', val_start)
                if val_end >= 0:
                    candidate = part[val_start:val_end]
                    if candidate and candidate != "$undefined" and not candidate.startswith("《"):
                        author = candidate

    # 提取推荐理由
    reason = ""
    idx_mt15 = part.find('mt-1.5')
    if idx_mt15 >= 0:
        idx_children = part.find('children', idx_mt15)
        if idx_children >= 0:
            idx_colon_bracket = part.find(':[', idx_children)
            if idx_colon_bracket >= 0:
                val_pos = idx_colon_bracket + 2
                bracket_end = part.find(']', val_pos)
                if bracket_end >= 0:
                    array_content = part[val_pos:bracket_end]
                    if '$undefined' in array_content:
                        first_comma = part.find(',', val_pos)
                        if first_comma >= 0 and first_comma < bracket_end:
                            second_quote = part.find('\\"', first_comma + 1)
                            if second_quote >= 0 and second_quote < bracket_end:
                                reason_start = second_quote + 2
                                reason_end = part.find('\\"', reason_start)
                                if reason_end >= 0 and reason_end <= bracket_end:
                                    reason = part[reason_start:reason_end]
                    elif '正文推荐' in array_content:
                        after_bracket = part.find(',\\"', bracket_end)
                        if after_bracket >= 0:
                            reason_start = after_bracket + 3
                            reason_end = part.find('\\"', reason_start)
                            if reason_end >= 0:
                                reason = part[reason_start:reason_end]
                else:
                    val_start = idx_colon_bracket + 1
                    if part[val_start:val_start+2] == '\\"':
                        reason_start = val_start + 2
                        reason_end = part.find('\\"', reason_start)
                        if reason_end >= 0:
                            r = part[reason_start:reason_end]
                            if r and r != '$undefined':
                                reason = r

    # 封面图
    cover_url = ""
    idx_src = part.find('src\\":\\"')
    if idx_src >= 0:
        src_start = idx_src + 8
        src_end = part.find('\\"', src_start)
        if src_end >= 0:
            img_path = part[src_start:src_end]
            if img_path.startswith('/') and any(ext in img_path for ext in ['.webp', '.png', '.jpg', '.jpeg']):
                cover_url = "https://www.jinjiancheng.com" + img_path

    # 详情链接
    detail_link = ""
    idx_href = part.find('href\\":\\"/go/book-')
    if idx_href >= 0:
        href_start = idx_href + 10
        href_end = part.find('\\"', href_start)
        if href_end >= 0:
            detail_link = "https://www.jinjiancheng.com" + part[href_start:href_end]

    return {
        "title": title,
        "author": author,
        "reason": html_lib.unescape(reason),
        "cover_url": cover_url,
        "detail_link": detail_link
    }


def extract_all_books_from_payload(payload):
    """
    从单个 payload 中提取所有书本信息。
    支持 li 和 h3 两种书本格式。
    """
    books = []

    # 尝试 li 格式分割
    li_parts = re.split(r'\[\\"\$\\",\\"li\\",\\"《', payload)
    if len(li_parts) > 1:
        for part in li_parts[1:]:
            info = _extract_book_info(part)
            if info:
                books.append(info)
        return books

    # 尝试 h3 格式
    h3_positions = []
    pos = 0
    while True:
        pos = payload.find('\\"h3\\"', pos)
        if pos == -1:
            break
        h3_positions.append(pos)
        pos += 1

    if not h3_positions:
        return books

    # 按 h3 位置分割 payload
    for i, hp in enumerate(h3_positions):
        start = hp
        end = h3_positions[i + 1] if i + 1 < len(h3_positions) else len(payload)
        part = payload[start:end]
        info = _extract_book_info(part)
        if info:
            books.append(info)

    return books


def parse_booklist(html_text):
    """
    解析书单数据。
    策略：
    1. 识别 section payload 和书本 payload
    2. 根据书本 ID（十六进制编号）确定分类归属
    3. section 内联书本直接归属对应分类
    """
    pattern = re.compile(r'self\.__next_f\.push\(\[1,"(.*?)"\]\)', re.DOTALL)
    payloads = pattern.findall(html_text)

    # 第一步：收集所有 section 分类
    sections = []  # [(category, [book_info])]
    section_categories = []  # 按顺序的分类名

    for payload in payloads:
        sec_match = re.search(r'\\"\$\\",\\"section\\",\\"([^\\]+)\\"', payload)
        if sec_match:
            category = sec_match.group(1)
            section_categories.append(category)
            # 提取该 section 中的内联书本
            inline_books = extract_all_books_from_payload(payload)
            sections.append((category, inline_books))

    # 第二步：收集所有独立书本 payload（非 section）及其 ID
    book_entries = []  # [(book_id, book_info)]
    inline_titles = set()
    for cat, books in sections:
        for b in books:
            inline_titles.add(b["title"])

    for payload in payloads:
        # 跳过 section payload
        if re.search(r'\\"\$\\",\\"section\\",\\"', payload):
            continue

        # 提取书本 ID
        id_match = re.match(r'^(\w+):\[', payload)
        if not id_match:
            continue
        book_id = id_match.group(1)

        # 提取书本信息
        extracted = extract_all_books_from_payload(payload)
        for info in extracted:
            if info["title"] not in inline_titles:
                book_entries.append((book_id, info))

    # 第三步：根据书本 ID 分配分类
    # 书本 ID 是十六进制，按已知范围分配
    # 投资/金融类: 1c-27, 逻辑学类: 36, 经典类: 37-44, 科学类: 45-46, 亲子类: 47+
    def hex_val(bid):
        try:
            return int(bid, 16)
        except ValueError:
            return 99999

    book_entries.sort(key=lambda x: hex_val(x[0]))

    if len(section_categories) >= 5:
        cat_invest = section_categories[0]
        cat_logic = section_categories[1]
        cat_classic = section_categories[2]
        cat_science = section_categories[3]
        cat_kids = section_categories[4]
    else:
        cat_invest = "投资 / 金融类"
        cat_logic = "逻辑学类"
        cat_classic = "经典 / 历史 / 人文类"
        cat_science = "科学 / 思维类"
        cat_kids = "亲子 / 青少年类"

    id_category_map = {}
    # 投资/金融类: 1c (28) 到 27 (39)
    for i in range(0x1c, 0x28):
        id_category_map[format(i, 'x')] = cat_invest
    # 逻辑学类: 36 (54)
    id_category_map['36'] = cat_logic
    # 经典/历史/人文类: 37 (55) 到 44 (68)
    for i in range(0x37, 0x45):
        id_category_map[format(i, 'x')] = cat_classic
    # 科学/思维类: 45 (69) 到 46 (70)
    for i in range(0x45, 0x47):
        id_category_map[format(i, 'x')] = cat_science
    # 亲子/青少年类: 47 (71) 以上
    id_category_map['47'] = cat_kids
    id_category_map['48'] = cat_kids

    # 第四步：组装最终结果
    # 先用字典存储，独立书本信息可以覆盖 section 内联信息
    books_dict = {}  # title -> {category, ...}

    # 先添加 section 内联书本
    for category, inline_books in sections:
        for info in inline_books:
            books_dict[info["title"]] = {"category": category, **info}

    # 再添加/覆盖独立书本（独立书本信息通常更完整）
    for book_id, info in book_entries:
        category = id_category_map.get(book_id, "未分类")
        # 如果独立书本有推荐理由或作者，覆盖内联版本
        if info["title"] in books_dict:
            existing = books_dict[info["title"]]
            if info["author"]:
                existing["author"] = info["author"]
            if info["reason"]:
                existing["reason"] = info["reason"]
            if info["cover_url"]:
                existing["cover_url"] = info["cover_url"]
            if info["detail_link"]:
                existing["detail_link"] = info["detail_link"]
            # 用独立书本的分类（基于 ID 映射，更准确）
            existing["category"] = category
        else:
            books_dict[info["title"]] = {"category": category, **info}

    books = list(books_dict.values())
    return books


def create_excel(books, output_path):
    """创建格式化的 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "金渐成推荐书单"

    hdr_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cat_font = Font(name="微软雅黑", size=11, bold=True, color="2F5496")
    cat_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    title_font = Font(name="微软雅黑", size=11, bold=True)
    normal_font = Font(name="微软雅黑", size=10)
    link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")

    c_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    l_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    bdr = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7")
    )

    headers = ["序号", "分类", "书名", "作者", "推荐理由", "封面图URL", "详情链接"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = bdr

    widths = [6, 22, 38, 18, 68, 55, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    row = 2
    for idx, book in enumerate(books, 1):
        vals = [idx, book["category"], book["title"], book["author"],
                book["reason"], book["cover_url"], book["detail_link"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = bdr
            if col == 1:
                c.font = normal_font; c.alignment = c_align
            elif col == 2:
                c.font = cat_font; c.fill = cat_fill; c.alignment = c_align
            elif col == 3:
                c.font = title_font; c.alignment = l_align
            elif col in (6, 7):
                c.font = link_font; c.alignment = l_align
                if val:
                    c.hyperlink = val
            else:
                c.font = normal_font; c.alignment = l_align
        row += 1

    # 汇总区
    row += 1
    sfill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    cats = {}
    for b in books:
        cats[b["category"]] = cats.get(b["category"], 0) + 1

    for col in range(1, 8):
        c = ws.cell(row=row, column=col, value="")
        c.fill = sfill; c.border = bdr
    c = ws.cell(row=row, column=2, value="分类汇总")
    c.font = Font(name="微软雅黑", size=11, bold=True, color="BF8F00")
    c.fill = sfill; c.alignment = c_align
    row += 1

    cat_order = ["投资 / 金融类", "逻辑学类", "经典 / 历史 / 人文类",
                 "科学 / 思维类", "亲子 / 青少年类"]
    for cat in cat_order:
        if cat in cats:
            for col in range(1, 8):
                c = ws.cell(row=row, column=col, value="")
                c.fill = sfill; c.border = bdr
            c = ws.cell(row=row, column=2, value=cat)
            c.font = Font(name="微软雅黑", size=10, bold=True); c.alignment = c_align
            c = ws.cell(row=row, column=3, value=f"{cats[cat]} 本")
            c.font = Font(name="微软雅黑", size=10); c.alignment = l_align
            row += 1

    for col in range(1, 8):
        c = ws.cell(row=row, column=col, value="")
        c.fill = sfill; c.border = bdr
    c = ws.cell(row=row, column=2, value="合计")
    c.font = Font(name="微软雅黑", size=10, bold=True, color="C00000"); c.alignment = c_align
    c = ws.cell(row=row, column=3, value=f"{len(books)} 本")
    c.font = Font(name="微软雅黑", size=10, bold=True, color="C00000"); c.alignment = l_align

    ws.row_dimensions[1].height = 32
    for r in range(2, row + 1):
        ws.row_dimensions[r].height = 28

    wb.save(output_path)
    print(f"\nExcel 已保存: {output_path}")
    print(f"共 {len(books)} 本书，{len(cats)} 个分类")


def main():
    print("=" * 60)
    print("  金渐成推荐书单爬虫")
    print("  https://www.jinjiancheng.com/booklist")
    print("=" * 60)

    print("\n[1/3] 抓取页面...")
    html = fetch_booklist_html()
    print(f"      成功 (HTML: {len(html)} 字符)")

    print("\n[2/3] 解析书单...")
    books = parse_booklist(html)

    if not books:
        print("      未解析到书籍，请检查页面结构")
        return

    cat_order = [
        "投资 / 金融类", "逻辑学类", "经典 / 历史 / 人文类",
        "科学 / 思维类", "亲子 / 青少年类"
    ]
    books.sort(key=lambda b: cat_order.index(b["category"])
               if b["category"] in cat_order else 99)

    print(f"      共 {len(books)} 本书\n")
    for i, b in enumerate(books, 1):
        a = f" - {b['author']}" if b['author'] else ""
        print(f"  {i:2d}. [{b['category']}] {b['title']}{a}")

    print("\n[3/3] 生成 Excel...")
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "金渐成推荐书单.xlsx")
    create_excel(books, out)

    cats = {}
    for b in books:
        cats[b["category"]] = cats.get(b["category"], 0) + 1
    print("\n" + "=" * 60)
    print("各分类统计:")
    for cat in cat_order:
        if cat in cats:
            print(f"  {cat}: {cats[cat]} 本")
    for cat, cnt in cats.items():
        if cat not in cat_order:
            print(f"  {cat}: {cnt} 本")
    print(f"  {'─' * 20}")
    print(f"  总计: {len(books)} 本")
    print("=" * 60)


if __name__ == "__main__":
    main()
